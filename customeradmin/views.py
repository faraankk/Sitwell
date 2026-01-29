from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth import login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.core.paginator import Paginator
from django.db.models import Q, Max, Sum, F
from django.db import transaction
from django.utils import timezone
from django.contrib.sessions.models import Session
import logging
from authenticate.models import Order, Wallet
from django.contrib.admin.views.decorators import staff_member_required
from .forms import CustomAuthenticationForm, ProductForm, ProductImageFormSet, OrderStatusForm
from .models import Product, ProductImage, Category
from .utils import process_image
from authenticate.models import Order, OrderItem
import io, csv, openpyxl
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum, Value
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import make_aware, now
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from authenticate.models import Coupon, Order
from .forms import CouponForm
from django.urls import reverse
from django.views.decorators.cache import never_cache
from django.template.response import TemplateResponse
from django.views.decorators.http import require_POST
from .models import Banner
from .forms import BannerForm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
from datetime import datetime
import os
logger = logging.getLogger(__name__)
User = get_user_model()

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def login_to_account(request):
    if request.user.is_authenticated and request.user.is_superuser:
        print("User is authenticated and superuser, redirecting to admin_dashboard")
        return redirect("customeradmin:admin_dashboard")
    if request.method == 'POST':
        print("POST data received:", request.POST)  
        form = CustomAuthenticationForm(data=request.POST)
        print("Form is valid:", form.is_valid())  
        
        if form.is_valid():
            user = form.get_user()
            print("User found:", user.email, "Is superuser:", user.is_superuser)
            
            if not user.is_superuser:
                messages.error(request, 'Only admin users can log in here.')
                print("User is not a superuser")
                return render(request, 'admin_login.html', {'form': form})
            
            login(request, user)
            username = user.first_name.title() if user.first_name else user.username
            messages.success(request, f"Login Successful. Welcome, {username}!")
            print("Login successful, redirecting to admin_dashboard")
            return redirect("customeradmin:admin_dashboard")
        else:
            print("Form errors:", form.errors)  
            messages.error(request, 'Invalid username or password. Please try again.')
            print("Invalid form data")
            return render(request, 'admin_login.html', {'form': form})
    
    form = CustomAuthenticationForm()
    return render(request, 'admin_login.html', {'form': form})


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def admin_dashboard(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')

   
    total_products = Product.objects.count()
    published_products = Product.objects.filter(status='published').count()
    low_stock_products = Product.objects.filter(status='low-stock').count()
    draft_products = Product.objects.filter(status='draft').count()
    
    recent_orders = Order.objects.all().order_by('-created_at')[:5]
    
    today = timezone.now().date()
    
    today_orders = Order.objects.filter(created_at__date=today)
    today_sales = today_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    today_orders_count = today_orders.count()
    
    total_revenue = Order.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    
    User = get_user_model()
    total_users = User.objects.filter(is_superuser=False).count()
    
    from django.db.models import Count, F
    best_selling_products_qs = OrderItem.objects.filter(
        order__status__in=['paid', 'shipped', 'delivered', 'out-for-delivery'],
        product__isnull=False
    ).values(
        'product__id', 'product__name', 'product__category', 'product__brand'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_sold')[:10]
    
    best_selling_products = []
    for p in best_selling_products_qs:
        best_selling_products.append({
            'name': p['product__name'],
            'category': p['product__category'],
            'brand': p['product__brand'],
            'total_sold': p['total_sold'],
            'total_revenue': p['total_revenue'],
        })
    
    best_selling_categories_qs = OrderItem.objects.filter(
        order__status__in=['paid', 'shipped', 'delivered', 'out-for-delivery'],
        product__isnull=False
    ).values(
        'product__category'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_sold')[:10]
    
    category_display_map = {cat.name: cat.name for cat in Category.objects.filter(is_deleted=False)}
    best_selling_categories = []
    for cat in best_selling_categories_qs:
        cat_dict = dict(cat)
        cat_dict['category_name'] = category_display_map.get(cat['product__category'], cat['product__category'] or 'Unknown')
        best_selling_categories.append(cat_dict)
    
    best_selling_brands_qs = OrderItem.objects.filter(
        order__status__in=['paid', 'shipped', 'delivered', 'out-for-delivery'],
        product__isnull=False,
        product__brand__isnull=False
    ).exclude(
        product__brand=''
    ).values(
        'product__brand'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_sold')[:10]
    
    best_selling_brands = []
    for b in best_selling_brands_qs:
        best_selling_brands.append({
            'brand_name': b['product__brand'],
            'total_sold': b['total_sold'],
            'total_revenue': b['total_revenue'],
        })
    
    current_year = timezone.now().year
    monthly_sales = Order.objects.filter(
        created_at__year=current_year,
        status__in=['paid', 'shipped', 'delivered', 'out-for-delivery']
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('month')
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    chart_labels = month_names
    chart_data = [0] * 12
    chart_orders = [0] * 12
    
    for sale in monthly_sales:
        if sale['month']:
            month_idx = sale['month'].month - 1
            chart_data[month_idx] = float(sale['total'] or 0)
            chart_orders[month_idx] = sale['count']
    
    context = {
        'total_products': total_products,
        'published_products': published_products,
        'low_stock_products': low_stock_products,
        'draft_products': draft_products,
        'recent_orders': recent_orders,
        'today': today,
        'today_sales': today_sales,
        'today_orders_count': today_orders_count,
        'total_revenue': total_revenue,
        'total_users': total_users,
        'best_selling_products': best_selling_products,
        'best_selling_categories': best_selling_categories,
        'best_selling_brands': best_selling_brands,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'chart_orders': chart_orders,
        'current_year': current_year,
    }
    
    return render(request, 'admin_dashboard.html', context)


@login_required
def chart_data_api(request):
    """API endpoint for fetching chart data with different filters"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    filter_type = request.GET.get('filter', 'monthly')
    
    from django.db.models import Count
    from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
    
    today = timezone.now()
    
    if filter_type == 'daily':
        start_date = today - timedelta(days=30)
        sales = Order.objects.filter(
            created_at__gte=start_date,
            status__in=['paid', 'shipped', 'delivered', 'out-for-delivery']
        ).annotate(
            period=TruncDay('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
        
        labels = []
        data = []
        orders = []
        for i in range(30):
            day = start_date + timedelta(days=i)
            labels.append(day.strftime('%d %b'))
            sale = next((s for s in sales if s['period'] and s['period'].date() == day.date()), None)
            data.append(float(sale['total']) if sale else 0)
            orders.append(sale['count'] if sale else 0)
            
    elif filter_type == 'weekly':
       
        start_date = today - timedelta(weeks=12)
        sales = Order.objects.filter(
            created_at__gte=start_date,
            status__in=['paid', 'shipped', 'delivered', 'out-for-delivery']
        ).annotate(
            period=TruncWeek('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
        
        labels = []
        data = []
        orders = []
        sales_list = list(sales)
        for i in range(12):
            week_start = start_date + timedelta(weeks=i)
            labels.append(f"Week {i+1}")
            sale = next((s for s in sales_list if s['period'] and s['period'].date() <= week_start.date() + timedelta(days=6) and s['period'].date() >= week_start.date()), None)
            data.append(float(sale['total']) if sale else 0)
            orders.append(sale['count'] if sale else 0)
            
    elif filter_type == 'yearly':
       
        current_year = today.year
        sales = Order.objects.filter(
            created_at__year__gte=current_year - 4,
            status__in=['paid', 'shipped', 'delivered', 'out-for-delivery']
        ).annotate(
            period=TruncYear('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
        
        labels = [str(current_year - i) for i in range(4, -1, -1)]
        data = [0] * 5
        orders = [0] * 5
        for sale in sales:
            if sale['period']:
                year_idx = sale['period'].year - (current_year - 4)
                if 0 <= year_idx < 5:
                    data[year_idx] = float(sale['total'] or 0)
                    orders[year_idx] = sale['count']
    else:
        current_year = today.year
        sales = Order.objects.filter(
            created_at__year=current_year,
            status__in=['paid', 'shipped', 'delivered', 'out-for-delivery']
        ).annotate(
            period=TruncMonth('created_at')
        ).values('period').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('period')
        
        labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        data = [0] * 12
        orders = [0] * 12
        for sale in sales:
            if sale['period']:
                month_idx = sale['period'].month - 1
                data[month_idx] = float(sale['total'] or 0)
                orders[month_idx] = sale['count']
    
    return JsonResponse({
        'labels': labels,
        'data': data,
        'orders': orders,
        'filter': filter_type
    })



@login_required
def customer_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission.")
        return redirect('/')
    
    return render(request, 'customers/customer_list.html')


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def product_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    products = Product.objects.all().order_by('-created_at')
    
   
    search_query = request.GET.get('search', '').strip()
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    
    
    status_filter = request.GET.get('status', '')
    if status_filter == 'out-of-stock':
        products = products.filter(Q(status='out-of-stock') | Q(stock_quantity=0))
    elif status_filter and status_filter != 'all' and status_filter != '':
        products = products.filter(status=status_filter)
    
    
    print(f"Products count after filtering: {products.count()}")
    print(f"Status filter applied: {status_filter}")
    
    paginator = Paginator(products, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    all_products = Product.objects.all()
    published_count = all_products.filter(status='published').count()
    low_stock_count = all_products.filter(status='low-stock').count()
    out_of_stock_count = all_products.filter(Q(status='out-of-stock') | Q(stock_quantity=0)).count()
    
    return render(request, 'products/product_list.html', {
        'products': page_obj.object_list,
        'page_obj': page_obj,
        'search_query': search_query,
        'current_status': status_filter,
        'status_choices': Product.STATUS_CHOICES,
        'total_products': products.count(),
        'published_count': published_count,        
        'low_stock_count': low_stock_count,        
        'out_of_stock_count': out_of_stock_count, 
    })




@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def add_product(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    if request.method == 'POST':
        print("=== ADD PRODUCT WITH MULTIPLE IMAGES DEBUG ===")
        form = ProductForm(request.POST)
        
        
        images = request.FILES.getlist('images')
        print(f"Number of images received: {len(images)}")

        if request.method == 'POST':
            print("=== DEBUGGING FORM SUBMISSION ===")
            print(f"POST data keys: {list(request.POST.keys())}")
            print(f"FILES data keys: {list(request.FILES.keys())}")
            print(f"All FILES: {request.FILES}")
            print(f"Images from getlist: {request.FILES.getlist('images')}")
            print("=====================================")
    

        
        
        if len(images) < 3:
            messages.error(request, "Please upload at least 3 images for the product.")
            return render(request, 'products/add_product.html', {'form': form})
        
        if len(images) > 6:
            messages.error(request, "Maximum 6 images allowed per product.")
            return render(request, 'products/add_product.html', {'form': form})
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    product = form.save()
                    print(f"Product saved: {product.id} - {product.name}")
                    
                   
                    for index, image in enumerate(images):
                        
                        processed_image = process_image(image)
                        
                        
                        product_image = ProductImage(
                            product=product,
                            image=processed_image,
                            is_primary=(index == 0),  
                            order=index
                        )
                        product_image.save()
                        print(f"Image {index + 1} saved for product {product.name}")
                    
                    messages.success(request, f"Product '{product.name}' with {len(images)} images added successfully!")
                    return redirect("customeradmin:product-list")
                    
            except Exception as e:
                error_msg = f"Error saving product: {str(e)}"
                print(error_msg)
                logger.error(error_msg)
                messages.error(request, error_msg)
        else:
            
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
        
        print("===========================================")
    else:
        form = ProductForm()
    
    return render(request, 'products/add_product.html', {'form': form})


@login_required

def edit_product(request, product_id):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        
       
        new_images = request.FILES.getlist('images')
        existing_images_count = product.images.count()
        total_images = existing_images_count + len(new_images)
        
        print(f"=== EDIT PRODUCT DEBUG ===")
        print(f"Existing images: {existing_images_count}")
        print(f"New images: {len(new_images)}")
        print(f"Total images: {total_images}")
        
        
        if total_images < 3:
            messages.error(request, f"Product must have at least 3 images. Currently has {existing_images_count}. Please upload {3 - existing_images_count} more images.")
            existing_images = product.images.all().order_by('order')
            images_count = existing_images.count()
            images_remaining = 6 - images_count
            return render(request, 'products/edit_product.html', {
                'form': form, 
                'product': product,
                'existing_images': existing_images,
                'images_count': images_count,
                'images_remaining': images_remaining,
                'min_images_required': max(0, 3 - images_count),
            })
        
        if total_images > 6:
            messages.error(request, "Maximum 6 images allowed per product.")
            existing_images = product.images.all().order_by('order')
            images_count = existing_images.count()
            images_remaining = 6 - images_count
            return render(request, 'products/edit_product.html', {
                'form': form, 
                'product': product,
                'existing_images': existing_images,
                'images_count': images_count,
                'images_remaining': images_remaining,
                'min_images_required': max(0, 3 - images_count),
            })
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    
                    updated_product = form.save()
                    
                    
                    if new_images:
                        current_max_order = product.images.aggregate(max_order=Max('order'))['max_order'] or -1
                        
                        for index, image in enumerate(new_images):
                           
                            if image.size > 5 * 1024 * 1024:  
                                messages.warning(request, f"Image '{image.name}' is too large (max 5MB). Skipped.")
                                continue
                            
                            if not image.content_type.startswith('image/'):
                                messages.warning(request, f"'{image.name}' is not a valid image. Skipped.")
                                continue
                            
                            processed_image = process_image(image)
                            
                            product_image = ProductImage(
                                product=product,
                                image=processed_image,
                                is_primary=False,  
                                order=current_max_order + index + 1
                            )
                            product_image.save()
                            print(f"New image {index + 1} added to product {product.name}")
                    
                    messages.success(request, f"Product '{updated_product.name}' updated successfully!")
                    return redirect("customeradmin:product-list")
                    
            except Exception as e:
                error_msg = f"Error updating product: {str(e)}"
                print(error_msg)
                logger.error(error_msg)
                messages.error(request, error_msg)
        else:
            
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ProductForm(instance=product)
    
    existing_images = product.images.all().order_by('order')
    images_count = existing_images.count()
    images_remaining = 6 - images_count
    
    return render(request, 'products/edit_product.html', {
        'form': form, 
        'product': product,
        'existing_images': existing_images,
        'images_count': images_count,
        'images_remaining': images_remaining,
        'min_images_required': max(0, 3 - images_count),
        'categories': Category.objects.filter(is_deleted=False, is_listed=True),
    })


@login_required
@require_POST
def delete_product_image(request, image_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        image = get_object_or_404(ProductImage, id=image_id)
        image.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error deleting product image: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def custom_logout(request):
    if request.method == 'POST':
        logout(request)
        messages.success(request, "You have been successfully logged out.")
    return redirect('customeradmin:login_to_account')




@login_required
@require_POST
def soft_delete_product(request, product_id):
    """Soft delete a product"""
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        product = get_object_or_404(Product.all_objects, id=product_id)
        
        if product.is_deleted:
            return JsonResponse({'success': False, 'error': 'Product is already deleted'})
        
        deleted_by = request.user.email if hasattr(request.user, 'email') else request.user.username
        product.soft_delete(deleted_by=deleted_by)
        
        return JsonResponse({
            'success': True,
            'message': f"Product '{product.name}' deleted successfully"
        })
        
    except Exception as e:
        logger.error(f"Error soft deleting product: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def restore_product(request, product_id):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission.")
        return redirect('/')
    
    try:
        product = get_object_or_404(Product.all_objects, id=product_id)
        
        if not product.is_deleted:
            messages.warning(request, f"Product '{product.name}' is not deleted.")
            return redirect("customeradmin:product-list")
        
        product.restore()
        messages.success(request, f"Product '{product.name}' has been restored successfully.")
        
    except Exception as e:
        logger.error(f"Error restoring product: {str(e)}")
        messages.error(request, f"Error restoring product: {str(e)}")
    
    return redirect("customeradmin:product-list")

@login_required
def deleted_products_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    deleted_products = Product.all_objects.filter(is_deleted=True).order_by('-deleted_at')
    
    paginator = Paginator(deleted_products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'products/deleted_products.html', {
        'products': page_obj,
        'page_obj': page_obj,
    })

@login_required
@require_POST
def delete_single_image(request, image_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        image = get_object_or_404(ProductImage, id=image_id)
        product = image.product
        
        remaining_images = product.images.exclude(id=image_id).count()
        if remaining_images < 3:
            return JsonResponse({
                'success': False, 
                'error': 'Cannot delete image. Product must have at least 3 images.'
            })
        
        image.delete()
        return JsonResponse({'success': True})
        
    except Exception as e:
        logger.error(f"Error deleting image: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})
    

@login_required
def category_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    categories = Category.objects.all().order_by('-created_at')
    
    search_query = request.GET.get('search', '').strip()
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if request.GET.get('clear'):
        return redirect('category-list')
    
    all_categories = Category.objects.filter(is_deleted=False)
    listed_count = all_categories.filter(is_listed=True).count()
    unlisted_count = all_categories.filter(is_listed=False).count()
    
    from django.utils import timezone
    current_month = timezone.now().month
    current_year = timezone.now().year
    new_this_month = all_categories.filter(
        created_at__month=current_month,
        created_at__year=current_year
    ).count()
    
    paginator = Paginator(categories, 5)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'category/category.html', {
        'categories': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'listed_count': listed_count,
        'unlisted_count': unlisted_count,
        'new_this_month': new_this_month,
    })


@login_required
@require_POST
def add_category(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        name = request.POST.get('name')
        is_listed = request.POST.get('is_listed') == 'true'
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})
        
        if Category.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'error': 'Category already exists'})
        
        category = Category.objects.create(
            name=name,
            is_listed=is_listed
        )
        
        return JsonResponse({'success': True, 'id': category.id, 'message': 'Category created successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def edit_category(request, category_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        category = get_object_or_404(Category, id=category_id)
        
        name = request.POST.get('name')
        is_listed = request.POST.get('is_listed') == 'true'
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})
        
        category.name = name
        category.is_listed = is_listed
        category.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def soft_delete_category(request, category_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        category = get_object_or_404(Category.all_objects, id=category_id)
        if category.is_deleted:
            return JsonResponse({'success': False, 'error': 'Category is already deleted'})
        
        deleted_by = request.user.email if hasattr(request.user, 'email') else request.user.username
        category.soft_delete(deleted_by=deleted_by)
        
        return JsonResponse({
            'success': True,
            'message': f"Category '{category.name}' deleted successfully"
        })
    except Exception as e:
        logger.error(f"Error soft deleting category: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})



@login_required
def deleted_categories_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')

    deleted_categories = Category.all_objects.filter(is_deleted=True).order_by('-deleted_at')

    paginator = Paginator(deleted_categories, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'category/deleted_categories.html', {
        'categories': page_obj,
        'page_obj': page_obj,
    })


@login_required
@require_POST 
def toggle_category_listed(request, category_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        category = get_object_or_404(Category, id=category_id)
        category.is_listed = not category.is_listed
        category.save()
        return JsonResponse({'success': True, 'is_listed': category.is_listed})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

@login_required
def deleted_categories_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')
    
    deleted_categories = Category.all_objects.filter(is_deleted=True).order_by('-deleted_at')
    
    paginator = Paginator(deleted_categories, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'category/deleted_categories.html', {
        'categories': page_obj,
        'page_obj': page_obj,
    })



@login_required
@require_POST
def restore_category(request, category_id):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission.")
        return redirect('/')

    category = get_object_or_404(Category.all_objects, id=category_id)

    if not category.is_deleted:
        messages.warning(
            request,
            f"Category '{category.name}' is not deleted."
        )
        return redirect("customeradmin:deleted-categories")

    try:
        category.restore()
        messages.success(
            request,
            f"Category '{category.name}' restored successfully."
        )
    except Exception as e:
        logger.error(f"Error restoring category: {e}")
        messages.error(request, "Error restoring category.")

    return redirect("customeradmin:deleted-categories")




User = get_user_model()

@login_required
def user_management_view(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to access this page.")
        return redirect('/')
    
    users = User.objects.filter(is_superuser=False).order_by('-created_at')
    
    search_query = request.GET.get('search', '').strip()
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    if request.GET.get('clear'):
        return redirect('user-management')
    
    paginator = Paginator(users, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    all_users = User.objects.filter(is_superuser=False)
    active_count = all_users.filter(is_blocked=False).count()
    blocked_count = all_users.filter(is_blocked=True).count()
    new_today = all_users.filter(created_at__date=timezone.now().date()).count()
    
    return render(request, 'User/user_management.html', {  
        'users': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'active_count': active_count,
        'blocked_count': blocked_count,
        'new_today': new_today,
    })

@login_required
@require_POST
def block_user(request, user_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        if user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Cannot block superuser'})
        
        if user.is_blocked:
            return JsonResponse({'success': False, 'error': 'User is already blocked'})
        
        
        user.block_user(blocked_by=request.user.email)
        
        
        sessions = Session.objects.all()
        for session in sessions:
            try:
                data = session.get_decoded()
                if str(user.id) == data.get('_auth_user_id'):
                    session.delete()
            except Exception:  # ✅ Better - only catches standard exceptions
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'User {user.email} has been blocked and logged out successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def unblock_user(request, user_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        if not user.is_blocked:
            return JsonResponse({'success': False, 'error': 'User is not blocked'})
        
        user.unblock_user()
        
        return JsonResponse({
            'success': True,
            'message': f'User {user.email} has been unblocked successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def order_list(request):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect("admindashboard")

    orders = Order.objects.all().order_by("-created_at")

    search = (request.GET.get("search") or "").strip()
    if search:
        orders = orders.filter(
            Q(order_number__icontains=search)
            | Q(user__email__icontains=search)
            | Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
        )

    status_filter = request.GET.get("status") or ""
    if status_filter and status_filter != "all":
        orders = orders.filter(status=status_filter)

    date_from = request.GET.get("from") or ""
    date_to = request.GET.get("to") or ""
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    sort = (request.GET.get("sort") or "").strip()
    sort_map = {
         "datedesc": "-created_at",
         "dateasc": "created_at",
         "totaldesc": "-total_amount",
         "totalasc": "total_amount",
    }
    if sort in sort_map:
        orders = orders.order_by(sort_map[sort])

    if request.GET.get("clear"):
        return redirect("customeradmin:order-list")

    paginator = Paginator(orders, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    if hasattr(Order, "Status") and getattr(Order.Status, "choices", None):
        status_tuple = Order.Status.choices
    elif hasattr(Order, "ORDERSTATUSCHOICES") and Order.ORDERSTATUSCHOICES:
        status_tuple = Order.ORDERSTATUSCHOICES
    else:
        status_tuple = (
            ("pending", "Pending"),
            ("paid", "Paid"),
            ("shipped", "Shipped"),
            ("out-for-delivery", "Out for Delivery"),
            ("delivered", "Delivered"),
            ("cancelled", "Cancelled"),
        )

    context = {
        "orders": page_obj.object_list,
        "pageobj": page_obj,
        "searchquery": search,
        "currentstatus": status_filter,
        "statuschoices": status_tuple,
        "sort": sort or "datedesc",
        "datefrom": date_from,
        "dateto": date_to,
        "totalorders": orders.count(),
    }
    return render(request, "orders/order_list.html", context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def order_detail(request, order_id: int):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect("admindashboard")

    order = get_object_or_404(Order, id=order_id)
    items = order.items.select_related("product").all().order_by("id")
    status_form = OrderStatusForm(order=order, initial={"status": order.status})
    context = {"order": order, "items": items, "statusform": status_form}
    return render(request, "orders/order_detail.html", context)


@login_required
@require_POST
@transaction.atomic
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def order_update_status(request, order_id: int):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("customeradmin:order-list")

    order = get_object_or_404(Order, id=order_id)
    form = OrderStatusForm(request.POST, order=order)
    if not form.is_valid():
        messages.error(request, "; ".join([str(v[0]) for v in form.errors.values()]))
        return redirect("order-detail", order_id=order.id)

    new_status = form.cleaned_data.get("status")
    if not new_status:
        messages.error(request, "Status is required.")
        return redirect("order-detail", order_id=order.id)

    old_status = order.status

    def deduct_for_paid():
        for item in order.items.select_related("product").all():
            product = item.product
            if not getattr(product, "manage_stock", True):
                continue
            
            if product.stock_quantity < item.quantity:
                raise ValueError(f"Insufficient stock for {product.name} (SKU {product.sku}).")
            product.stock_quantity -= item.quantity
            product.save() 

    def restock_for_cancel():
        for item in order.items.select_related("product").all():
            product = item.product
            if not getattr(product, "manage_stock", True):
                continue
            remaining = item.remaining_qty 
            if remaining > 0:
                product.stock_quantity += remaining
                item.cancelled_qty += remaining
                product.save()
                item.save(update_fields=["cancelled_qty", "updated_at"])

    def mark_shipped_full():
        for item in order.items.all():
            if item.remaining_qty > 0:
                item.mark_shipped_full()

    def mark_delivered_full():
        for item in order.items.all():
            if item.remaining_qty > 0:
                item.mark_delivered_full()

    try:
        if old_status == "pending" and new_status == "paid":
            deduct_for_paid()

        if new_status == "cancelled":
            restock_for_cancel()

        print(order.status)
        order.status =new_status
        order.updated_at = timezone.now() 
        order.save()
    except Exception as e:
        print(e)
        transaction.set_rollback(True)
        messages.error(request, f"Error updating order: {e}")
    return redirect("customeradmin:order-list")



@login_required
@require_POST
@transaction.atomic
@cache_control(no_cache=True, must_revalidate=True, no_store=True) 
def order_cancel(request, order_id: int):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("customeradmin:order-list")


    order = get_object_or_404(Order, id=order_id)
    if not order.can_be_cancelled():
        messages.error(request, "Order cannot be cancelled in its current status.")
        return redirect("order-detail", order_id=order.id)

    try:
        for item in order.items.select_related("product").all():
            product = item.product
            if not getattr(product, "manage_stock", True):
                continue
            remaining = item.remaining_qty
            if remaining > 0:
                product.stock_quantity += remaining
                item.cancelled_qty += remaining
                product.save()
                item.save(update_fields=["cancelled_qty", "updated_at"])

        order.set_status("cancelled")
        order.save()
        messages.success(request, f"Order {order.order_id} has been cancelled and stock restored.")
    except Exception as e:
        transaction.set_rollback(True)
        messages.error(request, f"Order {order.order_id} could not be cancelled: {e}")
    return redirect("order-detail", order_id=order.id)




@login_required
@require_POST
def verify_return_request(request, order_id):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("customeradmin:order-list")

    order = get_object_or_404(Order, order_number=order_id)

    if order.status != 'refund_pending':
        messages.error(request, "Order is not pending refund verification.")
        return redirect("customeradmin:order-list")

    order.status = 'refunded'
    order.payment_status = 'refunded'
    order.refund_processed_at = timezone.now()
    order.refund_processed_by = request.user.email
    order.save()

    wallet, _ = Wallet.objects.get_or_create(user=order.user)
    wallet.credit(
        order.total_amount,
        order,
        f"Refund for returned order {order.order_number}"
    )

    messages.success(
        request,
        f"Return verified & ₹{order.total_amount} refunded to customer wallet."
    )

    return redirect("customeradmin:order-list")



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def return_requests_list(request):
    """List all pending (or all) return requests for staff."""
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('/')

    qs = ReturnRequest.objects.select_related('order', 'processed_by')   

    status = request.GET.get('status', 'pending')
    if status != 'all':
        qs = qs.filter(status=status)

    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(order__order_number__icontains=search) |
            Q(order__user__email__icontains=search)
        )

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'return_requests': page_obj,
        'page_obj': page_obj,
        'current_status': status,
        'search_query': search,
        'status_choices': ReturnRequest.STATUS_CHOICES,

    }
    return render(request, 'orders/return_requests_list.html', context)



@login_required
@never_cache
def return_request_detail(request, return_id):
    return_request = get_object_or_404(ReturnRequest, id=return_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            try:
                return_request.status = 'approved'
                return_request.processed_at = timezone.now()
                return_request.processed_by = request.user
                return_request.save()
                
                wallet, created = Wallet.objects.get_or_create(
                    user=return_request.order.user,
                    defaults={'balance': 0}
                )
                
                wallet.balance += return_request.refund_amount
                wallet.save()
                
                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=return_request.refund_amount,
                    transaction_type='refund',
                    description=f"Refund for return request #{return_request.id}",
                    order=return_request.order,
                    status='completed'
                )
                
                return_request.order.status = 'refunded'
                return_request.order.save()
                
                messages.success(
                    request, 
                    f"Return request approved. ₹{return_request.refund_amount} refunded to customer's wallet."
                )
                
                return redirect('customeradmin:return_requests_list')
                
            except Exception as e:
                messages.error(request, f"Error processing refund: {str(e)}")
                
        elif action == 'reject':
            return_request.status = 'rejected'
            return_request.processed_at = timezone.now()
            return_request.processed_by = request.user
            return_request.save()
            
            messages.info(request, "Return request rejected.")
            return redirect('customeradmin:return_requests_list')
    
    context = {
        'return_request': return_request,
        'order': return_request.order,
        'items': return_request.items.all(),
        'customer': return_request.order.user,
    }
    
    return TemplateResponse(
        request, 
        'customeradmin/templates/return_request_detail.html', 
        context
    )



@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coupon_list(request):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("/")
    
    qs = Coupon.objects.all().order_by("-created_at")
    
    search_query = request.GET.get('search', '').strip()
    if search_query:
        qs = qs.filter(
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    today = timezone.now().date()
    
    active_count = qs.filter(is_active=True, valid_to__gte=timezone.now()).count()
    
    expired_count = qs.filter(valid_to__lt=timezone.now()).count()
    
    total_usage = qs.aggregate(total=Sum("used_count"))["total"] or 0

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "coupons/coupon_list.html", {
        "coupons": page_obj,
        "page_obj": page_obj,
        "active_count": active_count,
        "expired_count": expired_count,
        "total_usage": total_usage,
        "search_query": search_query,
        "today": today,  
    })
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coupon_add(request):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("/")
    
    if request.method == "POST":
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon = form.save(commit=False)
            coupon.created_by = request.user
            coupon.save()
            messages.success(request, f"Coupon {coupon.code} created successfully.")
            return redirect("customeradmin:coupon-list")
    else:
        form = CouponForm()
    
    return render(request, "coupons/coupon_form.html", {"form": form})

@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coupon_edit(request, coupon_id):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("/")
    
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == "POST":
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            form.save()
            messages.success(request, f"Coupon {coupon.code} updated successfully.")
            return redirect("customeradmin:coupon-list")
    else:
        form = CouponForm(instance=coupon)
    
    return render(request, "coupons/coupon_form.html", {
        "form": form, 
        "coupon": coupon
    })



@login_required
@require_POST
def coupon_delete(request, coupon_id):
    if not request.user.is_superuser:
        return JsonResponse({"success": False, "error": "Permission denied"})
    coupon = get_object_or_404(Coupon, id=coupon_id)
    coupon.delete()
    return JsonResponse({"success": True})

def _sales_qs(start, end):
    """Paid orders between two aware datetimes."""
    return Order.objects.filter(
        created_at__gte=start,
        created_at__lt=end,
        payment_status="paid"
    )


def _aggregate(qs):
    gross = qs.aggregate(total=Sum("total_amount"))["total"] or Decimal("0")
    discount = qs.aggregate(total=Sum("discount_amount"))["total"] or Decimal("0")
    coupon = qs.aggregate(total=Sum("coupon_discount"))["total"] or Decimal("0")
    return {
        "gross": gross,
        "discount": discount,
        "coupon": coupon,
        "net": gross - discount - coupon,
        "count": qs.count(),
    }


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def sales_report(request):
    if not request.user.is_superuser:
        messages.error(request, "Permission denied.")
        return redirect("/")

    filter_type = request.GET.get("filter", "day")  
    today = now().replace(hour=0, minute=0, second=0, microsecond=0)

    if filter_type == "day":
        start, end = today, today + timedelta(days=1)
    elif filter_type == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=7)
    elif filter_type == "month":
        start = today.replace(day=1)
        end = (start + timedelta(days=32)).replace(day=1)
    elif filter_type == "year":  
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31) + timedelta(days=1)
    else: 
        try:
            start_str = request.GET.get("from")
            end_str = request.GET.get("to")
            start = make_aware(datetime.strptime(start_str, "%Y-%m-%d"))
            end = make_aware(datetime.strptime(end_str, "%Y-%m-%d")) + timedelta(days=1)
        except Exception:
            messages.error(request, "Invalid date range.")
            return redirect("customeradmin:sales-report")

    qs = _sales_qs(start, end)
    summary = _aggregate(qs)
    daily = (
        qs.annotate(period=TruncDate("created_at"))
        .values("period")
        .annotate(
            gross=Sum("total_amount"),
            discount=Sum("discount_amount"),
            coupon=Sum("coupon_discount"),
            count=Sum(Value(1)),
        )
        .annotate(
            net=F("gross") - F("discount") - F("coupon")
        )
        .order_by("period")
    )

    request.session["report_filters"] = {
        "start": start.isoformat(),
        "end": end.isoformat(),
    }

    context = {
        "summary": summary,
        "daily": daily,
        "filter_type": filter_type,
        "start": start.date(),
        "end": (end - timedelta(seconds=1)).date(),  
    }
    
    try:
        return render(request, "reports/sales_report.html", context)
    except Exception as e:
        print(f"Template error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def _pdf_response(data, title="Sales Report"):
    
    
    font_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'fonts',
        'DejaVuSans.ttf'
    )
    pdfmetrics.registerFont(TTFont('DejaVu', font_path))
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    elements = []
    styles = getSampleStyleSheet()
    
    company_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4A5568'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='DejaVu'
    )
    elements.append(Paragraph("SITWELL FURNITURE", company_style))
    
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='DejaVu'
    )
    elements.append(Paragraph(title.upper(), title_style))
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='DejaVu'
    )
    generated_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    elements.append(Paragraph(f"Generated on: {generated_date}", date_style))
    
    elements.append(Spacer(1, 20))
    
    total_orders = sum(row["count"] for row in data)
    total_gross = sum(row["gross"] or 0 for row in data)
    total_discount = sum(row["discount"] or 0 for row in data)
    total_coupon = sum(row["coupon"] or 0 for row in data)
    total_net = total_gross - total_discount - total_coupon
    
    summary_data = [
        ["SUMMARY", "", "", ""],
        ["Total Orders", str(total_orders), "Total Discounts", f"₹{total_discount:,.2f}"],
        ["Gross Sales", f"₹{total_gross:,.2f}", "Coupon Discounts", f"₹{total_coupon:,.2f}"],
        ["NET SALES", f"₹{total_net:,.2f}", "", ""],
    ]
    
    summary_table = Table(summary_data, colWidths=[120, 120, 120, 120])
    summary_table.setStyle(TableStyle([
       
        ('SPAN', (0, 0), (3, 0)),
        ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (3, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (3, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (3, 0), 'DejaVu'),
        ('FONTSIZE', (0, 0), (3, 0), 12),
        ('BOTTOMPADDING', (0, 0), (3, 0), 12),
        ('TOPPADDING', (0, 0), (3, 0), 12),
        
        ('BACKGROUND', (0, 1), (3, 2), colors.HexColor('#F7FAFC')),
        ('FONTNAME', (0, 1), (3, 2), 'DejaVu'),
        ('ALIGN', (1, 1), (1, 2), 'RIGHT'),
        ('ALIGN', (3, 1), (3, 2), 'RIGHT'),
        ('FONTSIZE', (0, 1), (3, 2), 10),
        ('TOPPADDING', (0, 1), (3, 2), 8),
        ('BOTTOMPADDING', (0, 1), (3, 2), 8),
        
        ('SPAN', (0, 3), (1, 3)),
        ('SPAN', (2, 3), (3, 3)),
        ('BACKGROUND', (0, 3), (1, 3), colors.HexColor('#48BB78')),
        ('TEXTCOLOR', (0, 3), (1, 3), colors.whitesmoke),
        ('FONTNAME', (0, 3), (1, 3), 'DejaVu'),
        ('FONTSIZE', (0, 3), (1, 3), 12),
        ('ALIGN', (0, 3), (1, 3), 'CENTER'),
        ('TOPPADDING', (0, 3), (1, 3), 10),
        ('BOTTOMPADDING', (0, 3), (1, 3), 10),
        
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#CBD5E0')),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    breakdown_style = ParagraphStyle(
        'BreakdownHeader',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#4A5568'),
        spaceAfter=10,
        fontName='DejaVu'
    )
    elements.append(Paragraph("DAILY BREAKDOWN", breakdown_style))
    
    table_data = [["Date", "Orders", "Gross Sales", "Discounts", "Coupons", "Net Sales"]]
    
    for row in data:
        net = (row["gross"] or 0) - (row["discount"] or 0) - (row["coupon"] or 0)
        table_data.append([
            row["period"].strftime("%d %b %Y") if hasattr(row["period"], 'strftime') else str(row["period"]),
            str(row["count"]),
            f"₹{row['gross']:,.2f}" if row['gross'] else "₹0.00",
            f"₹{row['discount']:,.2f}" if row['discount'] else "₹0.00",
            f"₹{row['coupon']:,.2f}" if row['coupon'] else "₹0.00",
            f"₹{net:,.2f}",
        ])
    
    t = Table(table_data, colWidths=[80, 50, 85, 85, 85, 95])
    t.setStyle(TableStyle([
        
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVu'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2D3748')),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVu'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 1), (-1, -1), 8),
        ('RIGHTPADDING', (0, 1), (-1, -1), 8),
        
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
        
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#CBD5E0')),
    ]))
    
    elements.append(t)
    
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
        fontName='DejaVu'
    )
    elements.append(Paragraph("--- End of Report ---", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    return response


def _excel_response(data, title="Sales Report"):
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    headers = ["Date", "Orders", "Gross", "Discount", "Coupon", "Net"]
    ws.append(headers)
    for row in data:
        net = row["gross"] - row["discount"] - row["coupon"]
        ws.append([row["period"], row["count"], row["gross"], row["discount"], row["coupon"], net])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'
    return response

@login_required
def sales_report_download(request, format):
    if not request.user.is_superuser:
        return HttpResponse("Forbidden", status=403)
    
    print("=== SALES REPORT DOWNLOAD DEBUG ===")
    print(f"Format: {format}")
    
    filter_type = request.GET.get("filter", "day")
    
    try:
        if request.GET.get("from") and request.GET.get("to"):
           
            start_str = request.GET.get("from")
            end_str = request.GET.get("to")
            start = make_aware(datetime.strptime(start_str, "%Y-%m-%d"))
            end = make_aware(datetime.strptime(end_str, "%Y-%m-%d")) + timedelta(days=1)
        else:
            filters = request.session.get("report_filters")
            if not filters:
                return HttpResponse("No report filters found", status=400)
            start = datetime.fromisoformat(filters["start"])
            end = datetime.fromisoformat(filters["end"])
    except Exception as e:
        print(f"Error parsing dates: {e}")
        import traceback
        traceback.print_exc()
        return HttpResponse("Invalid date parameters", status=400)
    
    qs = _sales_qs(start, end)
    daily = (
        qs.annotate(period=TruncDate("created_at"))
        .values("period")
        .annotate(
            gross=Sum("total_amount"),
            discount=Sum("discount_amount"),
            coupon=Sum("coupon_discount"),
            count=Sum(Value(1)),
        )
        .order_by("period")
    )
    
    if format == "pdf":
        return _pdf_response(daily)
    if format == "excel":
        return _excel_response(daily)
    return HttpResponse("Bad format", status=400)



@login_required
def coupon_export(request, format):
    if not request.user.is_superuser:
        return HttpResponse("Forbidden", status=403)
    coupons = Coupon.objects.all().order_by("-created_at")
    if format == "pdf":
        return _coupon_pdf(coupons)
    if format == "excel":
        return _coupon_excel(coupons)
    return HttpResponse("Bad format", status=400)



def _coupon_pdf(coupons):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Coupon Report", styles["Title"]))
    data = [["Code", "Discount", "Min Order", "Used", "Expiry"]]
    for c in coupons:
        if hasattr(c, 'discount_type'):
            discount_display = f"{c.discount_percent}{'%' if c.discount_type == 'percentage' else ''}"
        else:
            discount_display = f"{c.discount_percent}%"
        
        data.append([c.code, discount_display,
                     f"₹{c.min_order_amount}", c.used_count, c.valid_to.date()])
    t = Table(data, hAlign="CENTER")
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                           ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                           ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                           ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                           ("FONTSIZE", (0, 0), (-1, 0), 10),
                           ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                           ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                           ("GRID", (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Coupons.pdf"'
    return response


def _coupon_excel(coupons):
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coupons"
    headers = ["Code", "Description", "Discount Type", "Discount", "Min Order", "Usage Limit", "Used", "Expiry", "Active"]
    ws.append(headers)
    for c in coupons:
        description = getattr(c, 'description', '') or ''
        
        if hasattr(c, 'discount_type'):
            discount_type_display = c.get_discount_type_display() if hasattr(c, 'get_discount_type_display') else c.discount_type
        else:
            discount_type_display = 'Percentage'
        
        ws.append([c.code, description, discount_type_display, c.discount_percent,
                   c.min_order_amount, c.max_usage, c.used_count, c.valid_to.date(), c.is_active])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Coupons.xlsx"'
    return response


@login_required
def banner_list(request):
    """List all banners with filtering and pagination"""
    banners = Banner.objects.all().order_by('-created_at')
    
    position_filter = request.GET.get('position')
    status_filter = request.GET.get('status')
    
    if position_filter:
        banners = banners.filter(position=position_filter)
    if status_filter:
        banners = banners.filter(status=status_filter)
    
    
    paginator = Paginator(banners, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'position_choices': Banner.POSITION_CHOICES,
        'status_choices': Banner.STATUS_CHOICES,
        'position_filter': position_filter,
        'status_filter': status_filter,
    }
    return render(request, 'banner/banner_list.html', context)

@login_required
def banner_add(request):
    """Add new banner"""
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES)
        if form.is_valid():
            banner = form.save()
            messages.success(request, f'Banner "{banner.title}" created successfully!')
            return redirect('customeradmin:banner-list')
    else:
        form = BannerForm()
    
    context = {
        'form': form,
        'title': 'Add New Banner',
        'position_choices': Banner.POSITION_CHOICES,
    }
    return render(request, 'banner/banner_form.html', context)

@login_required
def banner_edit(request, banner_id):
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            messages.success(request, f'Banner "{banner.title}" updated successfully!')
            return redirect('customeradmin:banner-list')
    else:
        form = BannerForm(instance=banner)
    
    context = {
        'form': form,
        'title': 'Edit Banner',
        'banner': banner,
        'position_choices': Banner.POSITION_CHOICES,
    }
    return render(request, 'banner/banner_form.html', context)

@login_required
def banner_delete(request, banner_id):
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        banner_title = banner.title
        banner.delete()
        messages.success(request, f'Banner "{banner_title}" deleted successfully!')
        return redirect('customeradmin:banner-list')
    
    context = {
        'banner': banner,
    }
    return render(request, 'banner/banner_delete_confirm.html', context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def wallet_transactions_list(request):
    """List all wallet transactions with filtering and pagination"""
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to access this page.")
        return redirect('/')
    
    from authenticate.models import WalletTransaction
    
    transactions = WalletTransaction.objects.select_related('wallet__user', 'order').all()
    

    txn_type_filter = request.GET.get('type', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if txn_type_filter:
        transactions = transactions.filter(txn_type=txn_type_filter)
    
    if user_filter:
        transactions = transactions.filter(
            Q(wallet__user__email__icontains=user_filter) |
            Q(wallet__user__first_name__icontains=user_filter) |
            Q(wallet__user__last_name__icontains=user_filter)
        )
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            transactions = transactions.filter(created_at__date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            transactions = transactions.filter(created_at__date__lte=to_date)
        except ValueError:
            pass
    
    total_credits = transactions.filter(txn_type='credit').aggregate(total=Sum('amount'))['total'] or 0
    total_debits = transactions.filter(txn_type='debit').aggregate(total=Sum('amount'))['total'] or 0
    total_transactions = transactions.count()
    
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page', 1)
    transactions = paginator.get_page(page)
    
    context = {
        'transactions': transactions,
        'total_credits': total_credits,
        'total_debits': total_debits,
        'total_transactions': total_transactions,
        'txn_type_filter': txn_type_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'wallet/wallet_transactions_list.html', context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def wallet_transaction_detail(request, transaction_id):
    """Detailed view for a single wallet transaction"""
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to access this page.")
        return redirect('/')
    
    from authenticate.models import WalletTransaction
    
    transaction = get_object_or_404(
        WalletTransaction.objects.select_related('wallet__user', 'order'),
        id=transaction_id
    )
    
    source_info = {
        'type': 'Unknown',
        'description': transaction.note or 'No description available',
        'order': None,
    }
    
    if transaction.order:
        order = transaction.order
        source_info['order'] = order
        
        if 'refund' in transaction.note.lower() or 'return' in transaction.note.lower():
            source_info['type'] = 'Product Return Refund'
        elif 'cancel' in transaction.note.lower():
            source_info['type'] = 'Order Cancellation Refund'
        elif transaction.txn_type == 'debit':
            source_info['type'] = 'Wallet Payment'
        else:
            source_info['type'] = 'Order Related Credit'
    else:
        if transaction.txn_type == 'credit':
            source_info['type'] = 'Manual Credit / Bonus'
        else:
            source_info['type'] = 'Manual Debit'
    
    context = {
        'transaction': transaction,
        'user': transaction.wallet.user,
        'source_info': source_info,
    }
    
    return render(request, 'wallet/wallet_transaction_detail.html', context)
